import openpyxl
import re
from typing import List, Dict

# ─── Índices de columna (base 0) del formato RPS ─────────────────────────────
COL_NUMERO      = 1   # N° — clave para distinguir sedes reales de resúmenes
COL_APS         = 2
COL_RUTA        = 11
COL_MUNICIPIO   = 12
COL_INSTITUCION = 13
COL_SEDE        = 14

# Columna BOLSAS de cada tipo de carne
COL_BOLSAS = {
    'cerdo_1': 26,
    'cerdo_2': 34,
    'res':     42,
    'pollo_1': 50,
    'pollo_2': 58,
    'pollo_3': 66,
}

# Columna PESO (kg) de cada tipo de carne
COL_PESO = {
    'cerdo_1': 27,
    'cerdo_2': 35,
    'res':     43,
    'pollo_1': 51,
    'pollo_2': 59,
    'pollo_3': 67,
}

FILA_DATOS = 8   # Los datos empiezan en la fila 8 (base 1)
HOJA_DATOS = 'CÁRNICOS'  # Siempre leer esta hoja, nunca la activa


def _safe_int(val) -> int:
    if val is None:
        return 0
    try:
        return int(float(str(val).replace(',', '').strip()))
    except (ValueError, TypeError):
        return 0


def _safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return round(float(str(val).replace(',', '').strip()), 4)
    except (ValueError, TypeError):
        return 0.0


def _es_numero_valido(val) -> bool:
    """Retorna True solo si val es un entero positivo (N° de sede real)."""
    if val is None:
        return False
    try:
        return int(float(str(val).strip())) > 0
    except (ValueError, TypeError):
        return False


def procesar_excel(ruta_archivo: str) -> List[Dict]:
    """
    Lee la hoja CÁRNICOS del Excel RPS y extrae una sede por fila.

    Regla de filtrado:
    - Solo procesa filas donde col[1] (N°) contiene un entero positivo.
    - Esto descarta automáticamente filas de cabecera, subtotales por módulo,
      resúmenes por municipio y la fila TOTAL al final del archivo.
    """
    wb = openpyxl.load_workbook(ruta_archivo, read_only=True, data_only=True)

    # Usar siempre la hoja CÁRNICOS, no wb.active
    # (la hoja activa puede variar según quién guardó el archivo)
    if HOJA_DATOS in wb.sheetnames:
        ws = wb[HOJA_DATOS]
    else:
        ws = wb.active  # fallback por si cambia el nombre

    sedes = []
    orden = 1

    for row in ws.iter_rows(min_row=FILA_DATOS, values_only=True):

        # ── Filtro 1: la columna N° debe ser un entero positivo ──────────────
        # Las filas de resumen (módulos, municipios, totales) tienen N° = None
        # o un string como "MÓDULO". Las sedes reales siempre tienen 1, 2, 3…
        numero = row[COL_NUMERO] if len(row) > COL_NUMERO else None
        if not _es_numero_valido(numero):
            continue

        # ── Filtro 2: debe tener nombre de sede ──────────────────────────────
        sede_nombre = row[COL_SEDE] if len(row) > COL_SEDE else None
        if not sede_nombre or not str(sede_nombre).strip():
            continue

        nombre_str = str(sede_nombre).strip()

        bolsas = {
            tipo: _safe_int(row[col])
            for tipo, col in COL_BOLSAS.items()
            if len(row) > col
        }
        pesos = {
            tipo: _safe_float(row[col])
            for tipo, col in COL_PESO.items()
            if len(row) > col
        }

        total_bolsas = sum(bolsas.values())
        total_peso   = round(sum(pesos.values()), 4)

        sedes.append({
            'nombre':         nombre_str,
            'institucion':    str(row[COL_INSTITUCION]).strip() if len(row) > COL_INSTITUCION and row[COL_INSTITUCION] else '',
            'municipio':      str(row[COL_MUNICIPIO]).strip()   if len(row) > COL_MUNICIPIO  and row[COL_MUNICIPIO]  else '',
            'ruta':           str(row[COL_RUTA]).strip()        if len(row) > COL_RUTA        and row[COL_RUTA]       else '',
            'aps':            _safe_int(row[COL_APS])           if len(row) > COL_APS         else 0,
            'cuota_asignada': total_bolsas,
            'bolsas_cerdo_1': bolsas.get('cerdo_1', 0),
            'bolsas_cerdo_2': bolsas.get('cerdo_2', 0),
            'bolsas_res':     bolsas.get('res',     0),
            'bolsas_pollo_1': bolsas.get('pollo_1', 0),
            'bolsas_pollo_2': bolsas.get('pollo_2', 0),
            'bolsas_pollo_3': bolsas.get('pollo_3', 0),
            'peso_total_kg':  total_peso,
            'orden':          orden,
        })
        orden += 1

    wb.close()
    return sedes


def procesar_pdf(ruta_archivo: str) -> List[Dict]:
    """Extracción básica desde PDF — fallback cuando no hay Excel."""
    try:
        import pdfplumber
    except ImportError:
        return []

    sedes  = []
    orden  = 1
    patron = re.compile(r'(.+?)\s+(\d+)\s*$')

    with pdfplumber.open(ruta_archivo) as pdf:
        for page in pdf.pages:
            texto = page.extract_text() or ''
            for linea in texto.split('\n'):
                linea = linea.strip()
                match = patron.match(linea)
                if match:
                    nombre = match.group(1).strip()
                    cuota  = int(match.group(2))
                    if len(nombre) > 3 and cuota > 0:
                        sedes.append({
                            'nombre':         nombre,
                            'institucion':    '',
                            'municipio':      '',
                            'ruta':           '',
                            'aps':            0,
                            'cuota_asignada': cuota,
                            'bolsas_cerdo_1': 0,
                            'bolsas_cerdo_2': 0,
                            'bolsas_res':     0,
                            'bolsas_pollo_1': cuota,
                            'bolsas_pollo_2': 0,
                            'bolsas_pollo_3': 0,
                            'peso_total_kg':  0.0,
                            'orden':          orden,
                        })
                        orden += 1
    return sedes